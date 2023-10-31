import openpyxl
import torch
import os
import pickle
book = openpyxl.Workbook()

DIR = './_pickles/'
list = os.listdir(DIR) # dir is your directory path

for file_name in list:
    #ws = book.add_sheet(model)
    if 'new' in file_name:
        print(file_name)
        ws = book.create_sheet()
        ws.title=file_name
        row = 2
        column = 1
        with open(DIR+file_name, 'rb') as fr:
            input_pkl = pickle.load(fr)
            u = torch.unique(input_pkl, return_counts=True)
            for i in range(1,len(u[0])//1048576+2):
                ws.cell(row=1, column=2*i-1).value='value'
                ws.cell(row=1, column=2*i).value='appearance'
            for i in range(len(u[0])):
                if row > 1048576:
                    row -= 1048575
                    column += 2
                ws.cell(row, column).value=float(u[0][i])
                ws.cell(row, column+1).value=float(u[1][i])
                row += 1
        fr.close()
    book.save('Distribution' + '.xlsx')

